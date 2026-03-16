"""
HCR Split-Initiator Probe Designer
====================================
Input  : FASTA file  (--fasta gene.fasta)
         or raw sequence (--seq ATGC...)
Output : Excel workbook with two sheets
         1. Probe_Design  – full QC metrics
         2. Order_Sheet   – oligos ready to send to synthesis vendor
"""

import sys
import re
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── HCR B-series initiator sequences ──────────────────────────────────────────
INITIATORS = {
    "B1": {"P1": "GAGGAGGGCAGCAAACGG",   "P2": "GAAGAGTCTTCCTTTACG",
            "S1": "AA", "S2": "TA"},
    "B2": {"P1": "CCTCGTAAATCCTCATCA",   "P2": "ATCATCCAGTAAACCGCC",
            "S1": "AA", "S2": "AA"},
    "B3": {"P1": "GTCCCTGCCTCTATATCT",   "P2": "CCACTCAACTTTAACCCG",
            "S1": "TT", "S2": "TT"},
    "B4": {"P1": "CCTCAACCTACCTCCAAC",   "P2": "TCTCACCATATTCGCTTC",
            "S1": "AA", "S2": "AT"},
    "B5": {"P1": "CTCACTCCCAATCTCTAT",   "P2": "CTACCCTACAAATCCAAT",
            "S1": "AA", "S2": "AA"},
}

# ── FASTA parser (no BioPython required) ──────────────────────────────────────
def parse_fasta(text: str) -> dict:
    """Return {header: sequence} from raw FASTA text."""
    records = {}
    header, parts = None, []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                records[header] = "".join(parts)
            header = line[1:].split()[0]
            parts = []
        else:
            parts.append(line.upper().replace("U", "T"))
    if header is not None:
        records[header] = "".join(parts)
    return records

def read_sequence(fasta_path: str = None, raw_seq: str = None):
    """Return (gene_name, sequence_string)."""
    if fasta_path:
        with open(fasta_path) as fh:
            text = fh.read()
        records = parse_fasta(text)
        if not records:
            sys.exit("ERROR: No sequences found in FASTA file.")
        name, seq = next(iter(records.items()))
        return name, seq.upper().replace("U", "T")
    elif raw_seq:
        seq = re.sub(r"[^ACGTUacgtu]", "", raw_seq).upper().replace("U", "T")
        return "Input_sequence", seq
    else:
        sys.exit("ERROR: Provide --fasta or --seq.")

# ── Sequence utilities ─────────────────────────────────────────────────────────
_COMP = str.maketrans("ATGCatgc", "TACGtacg")

def reverse_complement(seq: str) -> str:
    return seq.translate(_COMP)[::-1]

def gc_content(seq: str) -> float:
    s = seq.upper()
    return (s.count("G") + s.count("C")) / len(s) * 100

def tm_basic(seq: str) -> float:
    """Wallace rule Tm (°C)."""
    s = seq.upper()
    return 2 * (s.count("A") + s.count("T")) + 4 * (s.count("G") + s.count("C"))

def hairpin_score(seq: str) -> int:
    rc = reverse_complement(seq)
    best = 0
    for i in range(len(seq) - 4):
        for j in range(i + 5, len(seq) + 1):
            if seq[i:j] in rc:
                best = max(best, j - i)
    return best

def dimer_score(seq1: str, seq2: str) -> int:
    rc = reverse_complement(seq2)
    best = 0
    for i in range(len(seq1) - 4):
        for j in range(i + 5, len(seq1) + 1):
            if seq1[i:j] in rc:
                best = max(best, j - i)
    return best

def secondary_structure_score(target: str) -> int:
    rc = reverse_complement(target)
    return sum(1 for i in range(len(target) - 5) if target[i:i+6] in rc)

def passes_filters(gc, hp1, hp2, dim, struct,
                   gc_min=35, gc_max=65,
                   hp_max=6, dim_max=6, struct_max=4) -> tuple:
    flags = []
    if not (gc_min <= gc <= gc_max):  flags.append(f"GC {gc:.1f}%")
    if hp1 > hp_max:                  flags.append(f"Hairpin-P1 {hp1}")
    if hp2 > hp_max:                  flags.append(f"Hairpin-P2 {hp2}")
    if dim > dim_max:                  flags.append(f"Dimer {dim}")
    if struct > struct_max:            flags.append(f"Structure {struct}")
    return (len(flags) == 0, "; ".join(flags) if flags else "PASS")

# ── Probe generation ───────────────────────────────────────────────────────────
def generate_probes(seq: str, arm_type: str, step: int = 25,
                    target_len: int = 52) -> pd.DataFrame:
    arm_type = arm_type.upper()
    if arm_type not in INITIATORS:
        sys.exit(f"ERROR: arm_type must be one of {list(INITIATORS.keys())}")
    if step < target_len:
        sys.exit(f"ERROR: TILING_STEP ({step}) is less than target length ({target_len}). Set TILING_STEP >= {target_len} to prevent overlapping targets.")

    ini   = INITIATORS[arm_type]
    P1, P2, S1, S2 = ini["P1"], ini["P2"], ini["S1"], ini["S2"]
    rows  = []

    for i in range(0, len(seq) - target_len + 1, step):
        target = seq[i : i + target_len]
        arm1   = target[:25]
        arm2   = target[-25:]

        # Probes bind the mRNA → arms are RC of target arms
        probe1 = P1 + S1 + reverse_complement(arm1)
        probe2 = reverse_complement(arm2) + S2 + P2

        gc      = gc_content(target)
        hp1     = hairpin_score(probe1)
        hp2     = hairpin_score(probe2)
        dim     = dimer_score(probe1, probe2)
        struct  = secondary_structure_score(target)
        tm1     = tm_basic(probe1)
        tm2     = tm_basic(probe2)
        ok, flag = passes_filters(gc, hp1, hp2, dim, struct)

        if not ok:
            continue

        rows.append({
            "Probe_#":            len(rows) + 1,
            "Status":             "PASS",
            "Fail_reason":        "",
            "Target_start":       i + 1,
            "Target_end":         i + target_len,
            "Target_sequence":    target,
            "Arm1_5to3":          arm1,
            "Arm2_5to3":          arm2,
            "Probe1_5to3":        probe1,
            "Probe2_5to3":        probe2,
            "Probe1_length":      len(probe1),
            "Probe2_length":      len(probe2),
            "GC_percent":         round(gc, 1),
            "Tm_P1_C":            round(tm1, 1),
            "Tm_P2_C":            round(tm2, 1),
            "Hairpin_P1":         hp1,
            "Hairpin_P2":         hp2,
            "Dimer_score":        dim,
            "Structure_score":    struct,
            "Initiator_set":      arm_type,
        })

    return pd.DataFrame(rows)

def rank_and_select(df: pd.DataFrame, max_probes: int = 30) -> pd.DataFrame:
    """Score each probe and return the top N non-overlapping by quality score."""
    if df.empty:
        return df
    # Lower is better for all penalty metrics
    # Score = deviation from ideal GC (50%) + hairpin + dimer + structure
    df = df.copy()
    df["_score"] = (
        abs(df["GC_percent"] - 50) * 0.5 +
        df["Hairpin_P1"] * 2 +
        df["Hairpin_P2"] * 2 +
        df["Dimer_score"] * 2 +
        df["Structure_score"] * 1
    )
    df = df.sort_values("_score").head(max_probes)
    df = df.sort_values("Target_start").reset_index(drop=True)
    df["Probe_#"] = range(1, len(df) + 1)
    df = df.drop(columns=["_score"])
    return df

# ── Excel export ───────────────────────────────────────────────────────────────
# Colour palette
C_TEAL      = "1A7A8A"
C_TEAL_LITE = "D0EDF1"
C_GREEN     = "1E8449"
C_GREEN_LT  = "D5F5E3"
C_RED       = "C0392B"
C_RED_LT    = "FADBD8"
C_ORANGE    = "D35400"
C_ORANGE_LT = "FDEBD0"
C_GREY      = "F2F4F4"
C_WHITE     = "FFFFFF"
C_DARK      = "1C2833"

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _cell_font(size=10, bold=False, color=C_DARK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# Spacer highlight colour
C_SPACER = "E74C3C"   # red

def _rich_probe(probe_seq: str, initiator_part: str, spacer: str, spacer_after: bool) -> CellRichText:
    """Return CellRichText with spacer in red, rest in dark colour.
    probe1 = P1 + S1 + arm_rc  →  spacer_after=False  (spacer follows P1)
    probe2 = arm_rc + S2 + P2  →  spacer_after=True   (spacer follows arm)
    """
    normal_font  = InlineFont(rFont="Arial", sz=10*2, color=C_DARK)
    spacer_font  = InlineFont(rFont="Arial", sz=10*2, color=C_SPACER)
    s_len = len(spacer)
    i_len = len(initiator_part)
    if not spacer_after:
        # P1 | S1 | arm_rc
        p1_part  = probe_seq[:i_len]
        s_part   = probe_seq[i_len:i_len+s_len]
        arm_part = probe_seq[i_len+s_len:]
        return CellRichText(
            TextBlock(normal_font, p1_part),
            TextBlock(spacer_font, s_part),
            TextBlock(normal_font, arm_part),
        )
    else:
        # arm_rc | S2 | P2
        arm_part = probe_seq[:len(probe_seq)-i_len-s_len]
        s_part   = probe_seq[len(probe_seq)-i_len-s_len:len(probe_seq)-i_len]
        p2_part  = probe_seq[-i_len:]
        return CellRichText(
            TextBlock(normal_font, arm_part),
            TextBlock(spacer_font, s_part),
            TextBlock(normal_font, p2_part),
        )

def write_design_sheet(ws, df_all: pd.DataFrame, gene_name: str, arm_type: str):
    df_pass = df_all.reset_index(drop=True)

    # ── Title banner ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    ws["A1"] = f"HCR Probe Design Report  ·  Gene: {gene_name}  ·  Initiator: {arm_type}"
    ws["A1"].font      = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill      = _fill(C_TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Summary bar ───────────────────────────────────────────────────────────
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Probes passing QC: {len(df_pass)}   |   Oligos to order: {len(df_pass)*2}"
    ws["A2"].font      = Font(name="Arial", size=10, italic=True, color=C_DARK)
    ws["A2"].fill      = _fill(C_TEAL_LITE)
    ws["A2"].alignment = _left()
    ws.row_dimensions[2].height = 18

    # ── Section header ─────────────────────────────────────────────────────────
    ws.merge_cells("A4:R4")
    ws["A4"] = f"✔  PASSING PROBES  ({len(df_pass)} probes ready for ordering)"
    ws["A4"].font      = Font(name="Arial", size=11, bold=True, color=C_WHITE)
    ws["A4"].fill      = _fill(C_GREEN)
    ws["A4"].alignment = _left()
    ws.row_dimensions[4].height = 22

    # ── Column headers ─────────────────────────────────────────────────────────
    HEADERS = [
        "#", "Start", "End", "Target Sequence (52 nt)",
        "Arm1 (25 nt)", "Arm2 (25 nt)",
        "Probe 1  (5'→3')", "Probe 2  (5'→3')",
        "P1 Len", "P2 Len",
        "GC %", "Tm P1 (°C)", "Tm P2 (°C)",
        "Hairpin P1", "Hairpin P2", "Dimer", "Structure",
        "Initiator"
    ]
    COL_WIDTHS = [
        5, 7, 7, 34,
        16, 16,
        42, 42,
        7, 7,
        8, 10, 10,
        10, 10, 8, 10,
        10
    ]

    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font      = _header_font()
        cell.fill      = _fill(C_DARK)
        cell.alignment = _center()
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[5].height = 20

    # ── Data rows ──────────────────────────────────────────────────────────────
    ini = INITIATORS[arm_type]
    for r_off, (_, rec) in enumerate(df_pass.iterrows()):
        row = 6 + r_off
        keys = [
            "Probe_#", "Target_start", "Target_end", "Target_sequence",
            "Arm1_5to3", "Arm2_5to3",
            "Probe1_5to3", "Probe2_5to3",
            "Probe1_length", "Probe2_length",
            "GC_percent", "Tm_P1_C", "Tm_P2_C",
            "Hairpin_P1", "Hairpin_P2", "Dimer_score", "Structure_score",
            "Initiator_set"
        ]
        for ci, key in enumerate(keys, start=1):
            val = rec[key]
            cell = ws.cell(row=row, column=ci)
            # Rich-text spacer colouring for probe sequence columns
            if key == "Probe1_5to3":
                cell.value = _rich_probe(val, ini["P1"], ini["S1"], spacer_after=False)
            elif key == "Probe2_5to3":
                cell.value = _rich_probe(val, ini["P2"], ini["S2"], spacer_after=True)
            else:
                cell.value = val
                cell.font  = _cell_font(bold=(key == "Probe_#"))
            cell.fill      = _fill(C_GREEN_LT)
            cell.border    = _thin_border()
            cell.alignment = _left() if ci > 3 else _center()

    ws.freeze_panes = ws["A6"]


def write_order_sheet(ws, df_all: pd.DataFrame, gene_name: str):
    df_pass = df_all.reset_index(drop=True)

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Oligo Ordering Sheet  ·  {gene_name}  ·  HCR Probes"
    ws["A1"].font      = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill      = _fill(C_TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = (f"{len(df_pass)} probe pairs  ·  {len(df_pass)*2} oligos total  ·  "
                f"All 5'→3', unmodified, standard desalting purification")
    ws["A2"].font      = Font(name="Arial", size=10, italic=True, color=C_DARK)
    ws["A2"].fill      = _fill(C_TEAL_LITE)
    ws["A2"].alignment = _left()
    ws.row_dimensions[2].height = 18

    # ── Column headers ─────────────────────────────────────────────────────────
    HEADS = ["Oligo Name", "Sequence (5'→3')", "Length (nt)",
             "GC %", "Tm (°C)", "Purification", "Notes"]
    WIDTHS = [22, 52, 12, 8, 10, 14, 30]

    for ci, (h, w) in enumerate(zip(HEADS, WIDTHS), start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = _header_font()
        cell.fill      = _fill(C_DARK)
        cell.alignment = _center()
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    r = 4
    for idx, rec in df_pass.iterrows():
        pair_num = idx + 1
        ini = INITIATORS[rec["Initiator_set"]]

        for probe_num, (seq_key, label) in enumerate(
                [("Probe1_5to3", "P1"), ("Probe2_5to3", "P2")], start=1):

            seq  = rec[seq_key]
            name = f"{gene_name}_{rec['Initiator_set']}_pair{pair_num:02d}_{label}"
            gc   = round(gc_content(seq), 1)
            tm   = round(tm_basic(seq), 1)
            note = (f"Target pos {rec['Target_start']}–{rec['Target_end']}  "
                    f"| Arm{probe_num}")

            # Build rich-text sequence with spacer highlighted
            if label == "P1":
                rich_seq = _rich_probe(seq, ini["P1"], ini["S1"], spacer_after=False)
            else:
                rich_seq = _rich_probe(seq, ini["P2"], ini["S2"], spacer_after=True)

            bg   = C_GREEN_LT if probe_num == 1 else C_GREY
            vals = [name, rich_seq, len(seq), gc, tm, "STD desalt", note]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=ci)
                cell.value     = v
                if ci != 2:     # rich text cell gets no plain font override
                    cell.font  = _cell_font(bold=(ci == 1))
                cell.fill      = _fill(bg)
                cell.border    = _thin_border()
                cell.alignment = _left() if ci in (1, 2, 7) else _center()
            r += 1

        # thin visual divider between pairs
        for ci in range(1, 8):
            ws.cell(row=r, column=ci).fill = _fill(C_WHITE)
        ws.row_dimensions[r].height = 4
        r += 1

    ws.freeze_panes = ws["A4"]


def write_info_sheet(ws, gene_name, arm_type, seq_len, step, df_all):
    ini = INITIATORS[arm_type]
    df_pass = df_all

    ws.merge_cells("A1:D1")
    ws["A1"] = "HCR Probe Designer – Run Summary"
    ws["A1"].font      = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill      = _fill(C_TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    rows_info = [
        ("Gene / Sequence name",  gene_name),
        ("Sequence length (nt)",   seq_len),
        ("Initiator set",          arm_type),
        ("P1 initiator",           ini["P1"]),
        ("P2 initiator",           ini["P2"]),
        ("Spacer S1",              ini["S1"]),
        ("Spacer S2",              ini["S2"]),
        ("Tiling step (nt)",       step),
        ("Target arm length (nt)", 25),
        ("Full target span (nt)",  52),
        ("Max probes cap",         30),
        ("Probes in output",        len(df_all)),
        ("Oligos to order",        len(df_all) * 2),
        ("GC filter",              "35 % – 65 %"),
        ("Hairpin filter",         "≤ 6 nt"),
        ("Dimer filter",           "≤ 6 nt"),
        ("Structure filter",       "≤ 4 matches"),
    ]
    for i, (k, v) in enumerate(rows_info, start=2):
        ws.cell(row=i, column=1, value=k).font  = Font(name="Arial", size=10,
                                                        bold=True, color=C_DARK)
        ws.cell(row=i, column=1).fill           = _fill(C_TEAL_LITE)
        ws.cell(row=i, column=2, value=v).font  = Font(name="Arial", size=10,
                                                        color=C_DARK)
        ws.cell(row=i, column=2).fill           = _fill(C_WHITE)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 30


def export_excel(df: pd.DataFrame, gene_name: str, arm_type: str,
                 seq_len: int, step: int, outfile: str):
    wb = Workbook()

    ws_info   = wb.active
    ws_info.title = "Run_Info"
    ws_design = wb.create_sheet("Probe_Design")
    ws_order  = wb.create_sheet("Order_Sheet")

    write_info_sheet(ws_info,   gene_name, arm_type, seq_len, step, df)
    write_design_sheet(ws_design, df, gene_name, arm_type)
    write_order_sheet(ws_order,  df, gene_name)

    wb.save(outfile)
    print(f"\n✔  Saved: {outfile}")
    print(f"   Probes : {len(df)}  ({len(df)*2} oligos ready to order)")




FASTA_SEQUENCE = """
>your_gene_name
PASTE_YOUR_FASTA_SEQUENCE_HERE
"""

INITIATOR_SET  = "B1"          # Choose: B1 / B2 / B3 / B4 / B5
TILING_STEP    = 52            # Nucleotides between probe start positions — set to 52 (= target length) to prevent overlaps
OUTPUT_FILE    = os.path.join(os.path.expanduser("~"), "Downloads", "HCR_probes.xlsx")


if __name__ == "__main__":
    records = parse_fasta(FASTA_SEQUENCE)
    if not records:
        sys.exit("ERROR: No sequence found. Check your FASTA paste above.")

    gene_name, seq = next(iter(records.items()))
    seq = seq.upper().replace("U", "T")

    print(f"Gene     : {gene_name}")
    print(f"Seq len  : {len(seq)} nt")
    print(f"Initiator: {INITIATOR_SET}   Step: {TILING_STEP} nt")

    df = generate_probes(seq, INITIATOR_SET, step=TILING_STEP)

    if df.empty:
        sys.exit("No probes passed QC filters. Try a different initiator or step size.")

    df = rank_and_select(df, max_probes=30)
    print(f"Top probes selected: {len(df)}")

    export_excel(df, gene_name, INITIATOR_SET, len(seq), TILING_STEP, OUTPUT_FILE)
